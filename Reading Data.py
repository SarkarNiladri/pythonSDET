import openpyxl

path = "C:/Users/Niladri/OneDrive/Desktop/sample.xlsx"

workbook = openpyxl.load_workbook(path)
sheet = workbook["Sheet1"]

row = sheet.max_row
column = sheet.max_column

#for row in sheet.iter_rows(values_only=True):
 #   print(row)

for r in range(1, row+1):
    for c in range(1,column+1):
        print(sheet.cell(r,c).value)
    print(" ")